import os
import glob
from openpyxl import load_workbook
from utils.logger import logger


class ExcelReader:
    """
    Excel 测试数据读取器

    支持：
    - 单个 Sheet 读取（read_data）
    - 单个文件所有 Sheet 读取（read_all_sheets）
    - 目录下所有 Excel 文件批量扫描（load_from_directory）

    Excel 表格列说明（第一行为表头）：
    | 用例名称 | 请求方法 | 请求路径 | 请求头(JSON) | 请求参数(JSON) | 预期状态码 | 预期响应字段(JSON) |

    每条数据会自动附加两个元数据字段：
    - _file: 来源文件名（不含路径）
    - _sheet: 来源 Sheet 名称
    """

    # 表头映射
    COLUMN_MAP = {
        "用例名称": "name",
        "请求方法": "method",
        "请求路径": "path",
        "请求头": "headers",
        "请求参数": "body",
        "预期状态码": "expected_status",
        "预期响应字段": "expected_fields",
    }

    def __init__(self, file_path: str, sheet_name: str = None):
        self.file_path = file_path
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel 文件不存在: {file_path}")
        self.workbook = load_workbook(file_path, read_only=True)
        self.sheet_name = sheet_name
        self.sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        logger.info(f"加载 Excel: {file_path}, Sheet: {self.sheet.title}")

    def _parse_rows(self, sheet) -> list[dict]:
        """解析单个 Sheet 的所有行，返回字典列表"""
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = rows[0]
        col_keys = [self.COLUMN_MAP.get(h, h) for h in headers]

        cases = []
        for row in rows[1:]:
            if not any(row):  # 跳过空行
                continue
            case = dict(zip(col_keys, row))
            if "expected_status" in case and case["expected_status"]:
                case["expected_status"] = int(case["expected_status"])
            cases.append(case)
        return cases

    def read_data(self) -> list[dict]:
        """读取当前 Sheet 的所有测试用例数据，返回字典列表"""
        cases = self._parse_rows(self.sheet)
        file_name = os.path.basename(self.file_path)
        for c in cases:
            c.setdefault("_file", file_name)
            c.setdefault("_sheet", self.sheet.title)
        logger.info(f"[{file_name} / {self.sheet.title}] 共读取 {len(cases)} 条测试用例")
        self.workbook.close()
        return cases

    def read_all_sheets(self) -> list[dict]:
        """
        读取当前 Excel 文件的所有 Sheet，返回所有用例的合并列表。
        每条数据附加 _file（文件名）和 _sheet（Sheet名）元数据。
        """
        file_name = os.path.basename(self.file_path)
        all_cases = []
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            cases = self._parse_rows(sheet)
            for c in cases:
                c["_file"] = file_name
                c["_sheet"] = sheet_name
            all_cases.extend(cases)
            logger.info(f"[{file_name} / {sheet_name}] 读取 {len(cases)} 条用例")
        logger.info(f"[{file_name}] 所有 Sheet 共计 {len(all_cases)} 条用例")
        self.workbook.close()
        return all_cases

    @staticmethod
    def load_from_directory(dir_path: str, pattern: str = "*.xlsx") -> list[dict]:
        """
        扫描指定目录下所有匹配的 Excel 文件，读取全部 Sheet，返回合并后的用例列表。

        Args:
            dir_path: test_data 目录路径
            pattern:  文件匹配模式，默认 *.xlsx

        Returns:
            所有 Excel 文件中所有 Sheet 的测试用例列表
        """
        if not os.path.isdir(dir_path):
            raise NotADirectoryError(f"目录不存在: {dir_path}")

        excel_files = sorted(glob.glob(os.path.join(dir_path, pattern)))
        if not excel_files:
            logger.warning(f"目录 {dir_path} 下未找到匹配 {pattern} 的文件")
            return []

        logger.info(f"扫描目录: {dir_path}，找到 {len(excel_files)} 个 Excel 文件: "
                    f"{[os.path.basename(f) for f in excel_files]}")

        all_cases = []
        for file_path in excel_files:
            reader = ExcelReader(file_path)
            all_cases.extend(reader.read_all_sheets())

        logger.info(f"目录 {dir_path} 下共加载 {len(all_cases)} 条测试用例")
        return all_cases

    @staticmethod
    def parse_json_field(field: str) -> dict | None:
        """将 JSON 字符串字段解析为字典"""
        import json
        if not field or str(field).strip() == "":
            return None
        try:
            return json.loads(str(field))
        except json.JSONDecodeError as e:
            logger.warning(f"JSON 解析失败: {field}, 错误: {e}")
            return None
